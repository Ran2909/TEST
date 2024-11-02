import os
import openpyxl

# 指定文件夹路径和记录文件名字的xlsx文件路径
folder_path =r"C:\Users\29098\Desktop\学业\学通社\无课表"
record_file_path = r"C:\Users\29098\Desktop\学业\学通社\全社无课表.xlsx"

# 创建或打开记录文件名字的xlsx文件
record_workbook = openpyxl.Workbook()
record_sheet = record_workbook.active

# 指定记录范围 B2 到 H14
record_range = "B2:H14"

# 读取记录文件中的原有数据，并保存到字典中
previous_data = {}
for row in record_sheet[record_range]:
    for cell in row:
        previous_data[cell.coordinate] = cell.value


# 遍历指定文件夹中的所有xlsx文件
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)

        # 打开当前的xlsx文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 遍历B2到H14单元格
        for row in sheet.iter_rows(min_row=2, max_row=15, min_col=2, max_col=8):
            for cell in row:
                if cell.value is not None:
                    # 获取记录文件名字的行数和列数
                    row_num = cell.row
                    col_num = cell.column

                    # 获取文件名字(去掉文件后缀名)

                    file_name = os.path.splitext(filename)[0]

                    # 如果目标单元格为空，则添加文件名字；否则追加文件名字
                    if record_sheet.cell(row=row_num, column=col_num).value is None:
                        record_sheet.cell(row=row_num, column=col_num).value = f"{filename.split('.')[0]} ({cell.value})"
                    else:
                        # 检查之前的数据是否存在，如果存在则跳过
                        previous_value = previous_data.get(cell.coordinate)
                        if not previous_value or file_name not in previous_value:
                            if file_name == '摄影设计部' or file_name == '调研部':
                                record_sheet.cell(row=row_num, column=col_num).value += ", " + f"{''} ({cell.value})"

                            else:
                                record_sheet.cell(row=row_num, column=col_num).value += ", " + f"{filename.split('.')[0]} ({cell.value})"

        # 关闭当前的xlsx文件
        workbook.close()
print('wancheng')
# 保存记录文件名字的xlsx文件
record_workbook.save(record_file_path)
record_workbook.close()
