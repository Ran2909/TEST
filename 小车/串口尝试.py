import os
from openpyxl import load_workbook

# 指定文件夹路径和 Excel 文件路径
folder_path = 'D:/Documents'
excel_path = 'D:/output.xlsx'

# 获取指定文件夹中的所有文件名
files = os.listdir(folder_path)

# 存储符合条件的文件名
doc_names = []

for file_name in files:
    # 分离出文件名和后缀名
    name, ext = os.path.splitext(file_name)
    if ext == '.doc':
        doc_names.append(name)

# 打开 Excel 文件并获取第一个工作表
workbook = load_workbook(excel_path)
worksheet = workbook.worksheets[0]

# 添加文件名到第一列
for i, doc_name in enumerate(doc_names):
    cell = worksheet.cell(row=i+1, column=1)
    cell.value = doc_name

# 保存并关闭 Excel 文件
workbook.save(excel_path)
